import xlwt

wb = xlwt.Workbook()
sheet = wb.add_sheet('Movies')
sheet.write(0, 0, 'Html Movie Title')

import openpyxl as xl

wb = xl.load_workbook('movies.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
# Log statement to print the value of cell
print(cell.value)